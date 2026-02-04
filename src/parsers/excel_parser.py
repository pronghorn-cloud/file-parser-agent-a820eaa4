"""Excel spreadsheet parser using openpyxl."""

from pathlib import Path
from typing import Any, Dict, List, Optional
import logging

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .base import BaseDocumentParser, ParsedDocument

logger = logging.getLogger(__name__)


class ExcelParser(BaseDocumentParser):
    """Parser for Excel spreadsheets.
    
    Extracts cell data, sheets, formulas, and metadata from
    Excel files using the openpyxl library.
    """
    
    SUPPORTED_EXTENSIONS = ['.xlsx', '.xls']
    
    def __init__(self, file_path: Path, data_only: bool = True):
        """Initialize Excel parser.
        
        Args:
            file_path: Path to Excel file
            data_only: If True, read cell values instead of formulas
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")
        super().__init__(file_path)
        self._workbook = None
        self._data_only = data_only
        
        # Note: .xls files require xlrd
        if self.file_path.suffix.lower() == '.xls':
            logger.warning(".xls files may have limited support. Consider converting to .xlsx")
    
    @property
    def workbook(self):
        """Lazy-load Excel workbook."""
        if self._workbook is None:
            self._workbook = load_workbook(
                str(self.file_path), 
                data_only=self._data_only,
                read_only=False
            )
        return self._workbook
    
    def parse(self) -> ParsedDocument:
        """Parse Excel file and return structured content."""
        logger.info(f"Parsing Excel file: {self.file_path}")
        
        doc = ParsedDocument(
            filename=self.file_path.name,
            file_type='excel'
        )
        
        try:
            doc.metadata = self.extract_metadata()
            doc.content = {
                'sheet_count': len(self.workbook.sheetnames),
                'sheet_names': self.workbook.sheetnames,
                'sheets': self._extract_sheets()
            }
            doc.tables = self.extract_tables()
        except Exception as e:
            logger.error(f"Error parsing Excel file: {e}")
            doc.errors.append(str(e))
        
        return doc
    
    def extract_text(self) -> str:
        """Extract all text from Excel file."""
        text_parts = []
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            text_parts.append(f"=== {sheet_name} ===")
            
            for row in sheet.iter_rows():
                row_text = []
                for cell in row:
                    if cell.value is not None:
                        row_text.append(str(cell.value))
                if row_text:
                    text_parts.append("\t".join(row_text))
            
            text_parts.append("")  # Empty line between sheets
        
        return "\n".join(text_parts)
    
    def extract_metadata(self) -> Dict[str, Any]:
        """Extract Excel file metadata."""
        metadata = {}
        props = self.workbook.properties
        
        if props:
            prop_names = [
                'creator', 'title', 'description', 'subject',
                'keywords', 'category', 'created', 'modified',
                'lastModifiedBy', 'revision', 'version'
            ]
            
            for prop in prop_names:
                value = getattr(props, prop, None)
                if value:
                    if hasattr(value, 'isoformat'):
                        value = value.isoformat()
                    metadata[prop] = str(value)
        
        # Add workbook statistics
        metadata['sheet_count'] = len(self.workbook.sheetnames)
        metadata['sheet_names'] = self.workbook.sheetnames
        
        return metadata
    
    def _extract_sheets(self) -> List[Dict[str, Any]]:
        """Extract data from all sheets."""
        sheets = []
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            sheet_data = {
                'name': sheet_name,
                'dimensions': sheet.dimensions,
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'data': self._extract_sheet_data(sheet),
                'merged_cells': [str(r) for r in sheet.merged_cells.ranges]
            }
            sheets.append(sheet_data)
        
        return sheets
    
    def _extract_sheet_data(self, sheet, max_rows: Optional[int] = None) -> List[List[Any]]:
        """Extract cell data from a sheet.
        
        Args:
            sheet: openpyxl worksheet
            max_rows: Maximum rows to extract (None for all)
            
        Returns:
            2D list of cell values
        """
        data = []
        row_count = 0
        
        for row in sheet.iter_rows():
            if max_rows and row_count >= max_rows:
                break
            
            row_data = []
            for cell in row:
                value = cell.value
                # Convert special types to string
                if value is not None:
                    if hasattr(value, 'isoformat'):  # datetime
                        value = value.isoformat()
                    elif not isinstance(value, (str, int, float, bool)):
                        value = str(value)
                row_data.append(value)
            
            # Only add non-empty rows
            if any(v is not None for v in row_data):
                data.append(row_data)
                row_count += 1
        
        return data
    
    def extract_tables(self) -> List[Dict[str, Any]]:
        """Extract tables from Excel.
        
        In Excel, each sheet is essentially a table.
        This method returns sheet data in a table-friendly format.
        """
        tables = []
        for sheet_name in self.workbook.sheetnames:
            sheet = self.workbook[sheet_name]
            if sheet.max_row and sheet.max_column:
                table_data = {
                    'name': sheet_name,
                    'rows': sheet.max_row,
                    'columns': sheet.max_column,
                    'headers': [],
                    'data': []
                }
                
                # Assume first row is headers
                first_row = True
                for row in sheet.iter_rows():
                    row_values = [cell.value for cell in row]
                    if any(v is not None for v in row_values):
                        if first_row:
                            table_data['headers'] = row_values
                            first_row = False
                        else:
                            table_data['data'].append(row_values)
                
                tables.append(table_data)
        
        return tables
